from string import ascii_uppercase
from datetime import datetime as dt
from openpyxl import Workbook
from openpyxl.worksheet import table
from openpyxl.styles import Font, PatternFill, numbers
import scipy.optimize
from language import lang
from investment import Investment


# Reference:
# https://stackoverflow.com/questions/8919718/financial-python-library-that-has-xirr-and-xnpv-function/
def xnpv(rate, values, dates):
    """
    Similar to Excel's XNPV function.
    dates: list of dt objects
    """
    if rate <= -1.0:
        return float('inf')
    d0 = dates[0]  # dates in ASC order
    return sum(
        [
            vi / ((1.0 + rate) ** ((di - d0).days / 365.0))
            for vi, di in zip(values, dates)
        ]
    )


def xirr(values, dates):
    """
    Similar to Excel's XIRR function.
    dates: list of dt objects
    """
    try:
        return scipy.optimize.newton(lambda r: xnpv(r, values, dates), 0.0)
    except RuntimeError:
        return scipy.optimize.brentq(lambda r: xnpv(r, values, dates), -1.0, 1e10)


def interpolate_color(start_color, end_color, factor):
    return tuple(
        int(start + (end - start) * factor)
        for start, end in zip(start_color, end_color)
    )


def irr_color(irr):
    RED = (0xe0, 0x07, 0x07)
    GREEN = (0x66, 0xff, 0x00)

    if irr >= 0.04:
        color = GREEN
    elif irr <= 0.02:
        color = RED
    else:
        factor = (irr - 0.02) / 0.02
        color = interpolate_color(RED, GREEN, factor)

    return f'{color[0]:02x}{color[1]:02x}{color[2]:02x}'


def calculate_values_and_incomes(history, product):
    values = [h['capital'] for h in history]
    incomes = [h['capital'] for h in history if h['type'] != 'subscription']
    current_value = round(product['price'] * product['share'], 4)
    values.append(current_value)
    incomes.append(current_value)
    return values, incomes


def calculate_dates(history, product):
    dates = [dt.strptime(h['date'], '%Y-%m-%d') for h in history]
    current_date = product['last_update']
    dates.append(dt.strptime(current_date, '%Y-%m-%d'))
    return dates


def calc_financials(product):
    history = Investment('money.db').fetch_history(product['id'])
    # values include subscriptions but incomes not
    values, incomes = calculate_values_and_incomes(history, product)
    dates = calculate_dates(history, product)
    value = sum(incomes)
    profit = sum(values)
    irr = xirr(values, dates)
    return value, profit, irr


def generate_report_header(ws):
    headers = (lang['Code'], lang['Name'], lang['Due'],
               lang['Capital'], lang['Value'], lang['Profit'], lang['IRR'],
               lang['Baseline'], lang['StocksPct'], lang['BondsPct'])
    columns = list(ascii_uppercase)
    for i, h in enumerate(headers):
        ws[f'{columns[i]}1'] = h
        ws[f'{columns[i]}1'].style = 'Headline 2'


def generate_report(products):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Running Products'
    ws_ended = wb.create_sheet(title='Ended Products')

    generate_report_header(ws)
    generate_report_header(ws_ended)

    ws_row = populate_worksheet(ws,
                                [p for p in products
                                 if float(p['share']) > 0.0])
    ws_ended_row = populate_worksheet(ws_ended,
                                      [p for p in products
                                       if float(p['share'] == 0.0)])

    setup_worksheet(ws, ws_row)
    setup_worksheet(ws_ended, ws_ended_row)

    wb.save(f'Report_{dt.now().strftime("%Y-%m-%d")}.xlsx')


def populate_worksheet(ws, products):
    row = 0
    for p in products:
        value, profit, irr = calc_financials(p)
        fill_product_row(ws, row + 2, p, value, profit, irr)
        row += 1
    return row


def fill_product_row(ws, row, product, value, profit, irr):
    ws[f'A{row}'] = product['id']
    ws[f'B{row}'] = product['name']
    ws[f'C{row}'] = product['first_due']
    ws[f'D{row}'] = product['capital']
    ws[f'E{row}'] = value
    ws[f'F{row}'] = profit
    cell = f'G{row}'
    ws[cell] = irr
    ws[cell].font = Font(bold=True)
    ws[cell].fill = PatternFill(fill_type='solid', fgColor=irr_color(irr))
    ws[f'H{row}'] = product['baseline']
    ws[f'I{row}'] = product['stocks_pct']
    ws[f'J{row}'] = product['bonds_pct']


def setup_worksheet(ws, row_count):
    tab = table.Table(displayName=ws.title.replace(' ', ''), ref=f'A1:J{row_count + 1}')
    ws.add_table(tab)
    set_column_widths(ws)
    format_percentage_column(ws, 'G')


def set_column_widths(ws):
    column_widths = {
        'A': 20, 'B': 40, 'C': 20, 'D': 15, 'E': 15, 'F': 15, 'G': 10, 'H': 20, 'I': 10, 'J': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


def format_percentage_column(ws, col):
    for cell in ws[col]:
        cell.number_format = numbers.FORMAT_PERCENTAGE_00
